from django.http import Http404
from django.shortcuts import redirect
from io import BytesIO
from random import randint
import uuid
import base64
from django.core.files.base import ContentFile
from PIL import Image
from django.db.models import Q
from usercp.models import Role
from startup.models import StartUp, Referee
from usertracker.models import TheStatus, UserTracker, UserContent
from message.models import Notification
from usercp.models import User
from django.http import HttpResponse, HttpResponseRedirect, HttpResponseServerError
import xlwt
from urllib.parse import urlsplit
from django.contrib.contenttypes.models import ContentType
from django_user_agents.utils import get_user_agent
import logging
from django.views.defaults import server_error
from django.contrib.sites.shortcuts import get_current_site
from urllib.parse import urlsplit
from django.core.mail import EmailMultiAlternatives
from django.utils.html import strip_tags
from django.template.loader import render_to_string
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from persiantools.jdatetime import JalaliDate
from functools import wraps
from email.mime.image import MIMEImage
from .object import get_role_service

logger = logging.getLogger(__name__)


def get_client_ip(request):
    """get requested user ip address
    
    Arguments:
        request {REQUEST} -- requested ip
    
    Returns:
        REMOTE_IP -- returns ip address of requsted user
    """
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def premession(role, th_type1=None,  th_type2=None):
    """a premession decorator to determine access of requested user ro a path
    
    Arguments:
        role {STRING} -- the name field of Role Model to specify requested user role
    
    Keyword Arguments:
        th_type1 {STRING} -- specify user type(if is a manager or startup or karmand) (default: {None})
        th_type2 {STRING} -- specify user type(if is a manager or startup or karmand) (default: {None})
    
    Returns:
        FUNCTION -- returns the wraped function if the decorator validate
    """
    
    def function(function):
        @wraps(function)
        def wrapper(request, *args, **kwargs):
            if th_type1 is not None or th_type2 is not None:
                if request.user.user_type == th_type1 or request.user.user_type == th_type2 or request.user.is_admin:
                    if th_type2 == 'karmand' and request.user.can_see_startups == True:
                        return function(request, *args, **kwargs)

                    return function(request, *args, **kwargs)
                else:
                    if role:
                        r = Role.objects.filter(name__in=role)
                        check = any(
                            item in r for item in request.user.role.all())

                        if check is True:

                            return function(request, *args, **kwargs)
                        else:

                            return HttpResponseRedirect(request.META.get("HTTP_REFERER"))

                    return HttpResponseRedirect(request.META.get("HTTP_REFERER"))
        return wrapper
    return function


def logging_view(f):
    """a decorator to log the errors and inforamtion about server and views
    
    Arguments:
        f {FUNCTION} -- the wraped function
    
    Returns:
        FUNCTION -- wraped function if no error happes else it will return status of 500 
    """
    @wraps(f)
    def wrapper(request, *args, **kwargs):
        try:
            return f(request, *args, **kwargs)
        except Exception as e:
            if 'matches the given query' in (str(e)):
                raise Http404()
            logger.info(str(e))
            return server_error(request, template_name='500.html')
    return wrapper


def export_list_csv(the_model):
    """export an xlsx format file of informations depend on model type
    
    Arguments:
        the_model {STRING} -- specify from which model the information should be exported
    
    Returns:
        XLSX -- returns an excle file
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-{movies}.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
        movies=the_model,
    )
    workbook = Workbook()

    worksheet = workbook.active
    if the_model == 'ref_lead':
        worksheet.title = 'Leader'
        columns = ['نام کاربری', 'نام', 'نام خانوادگی', 'ایمیل', 'تلفن', 'نقش']
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        coach = get_role_service(name='coach')
        leader = get_role_service(name='leader')
        leaders_referees = User.objects.filter(Q(role=coach) | Q(role=leader))
        for lead_ref in leaders_referees:
            row_num += 1
            row = [lead_ref.username, lead_ref.first_name, lead_ref.last_name, lead_ref.email,
                   lead_ref.phone, 'راهبر' if leader in lead_ref.role.all() else 'مربی', ]
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
    if the_model == 'referee':
        worksheet.title = 'Referee'
        columns = ['نام و نام خانوادگی داور', 'استارت اپ ها']
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        referees = Referee.objects.filter(the_type='not_pres')
        for ref in referees:
            row_num += 1
            row = ["{a} {b}".format(
                a=ref.user.first_name, b=ref.user.last_name), ref.startup.title, ]
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
    if the_model == "startup":
        worksheet.title = 'StartUp'
        columns = ['استارت اپ', 'نام و نام خانوادگی داور']
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        refers = Referee.objects.filter(the_type='not_pres')
        starts = StartUp.objects.filter(referee__in=refers).distinct()
        for start in starts:
            row_num += 1
            row = [f"{start.title}", '-'.join(["{a} {b}".format(a=s.user.first_name, b=s.user.last_name)
                                               for s in start.referee.filter(the_type='not_pres')]), ]
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
        workbook.save(response)

    if the_model == 'the_startup':
        worksheet.title = 'startup'

        columns = ['نام استارتاپ', 'نام', 'نام خانوادگی', 'ایدی', 'تلفن', 'وضغیت استارتاپ']
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        leaders_referees = StartUp.objects.all()
        for lead_ref in leaders_referees:
            row_num += 1
            row = [lead_ref.title, lead_ref.owner.first_name, lead_ref.owner.last_name,
                   lead_ref.id, lead_ref.owner.phone, lead_ref.status]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
    if the_model == 'user':
        worksheet.title = 'user'
        columns = ['نام', 'نام خانوادگی', 'ایدی', 'تلفن']
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        referees = User.objects.all()
        for ref in referees:
            row_num += 1
            row = [ref.first_name, ref.last_name, ref.id, ref.phone]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
        workbook.save(response)
    return response


@logging_view
def status_of_user(tracker_info, status, name=None, content1=None, content2=None):
    """it will append some additional info to the latest requested user record in UserTracker table
    
    Arguments:
        tracker_info {DICT} -- a dictionary that contains requested user id an ip
        status {STRING} -- to specify detailed info
    
    Keyword Arguments:
        name {STRING} -- an status to user in choice field of TheStatus table (default: {None})
        content1 {STRING} -- additional info if available for UserContent table (default: {None})
        content2 {STRING} -- additional info if available for UserContent table (default: {None})
    
    Returns:
        OBJECT -- returns modified record
    """
    try:
        try:
            tracker = UserTracker.objects.filter(
                user_ip=get_client_ip(tracker_info['user_ip'])).order_by('-created_date').first()
        except:
            tracker = UserTracker.objects.filter(
                user__pk=tracker_info['user_id']).order_by('-created_date').first()

        tracker.status = status
        if name is not None:
            the_status = TheStatus.objects.get(name=name)
            tracker.the_status = the_status
        if not status and not name:
            tracker.delete()
        if content1 is not None:
            UserContent.objects.create(user_tracker=tracker, content=content1)
        if content2 is not None:
            UserContent.objects.create(user_tracker=tracker, content=content2)
        return tracker.save()
    except Exception as e:
        logger.info(str(e))



@logging_view
def send_email(track_info, **kwargs):
    """send email from server to determined email
    
    Arguments:
        protocol {SCHEME} -- the protocol of the uri (HTTP/HTTPS)
        current_site {STRING} -- the current domain
    
    Keyword Arguments:
        the_messages {STRING} -- a message that should be in the body of email (default: {None})
        the_messages2 {STRING} -- a message that should be in the body of email (default: {None})
        is_team {STRING} -- specify if the email should be sent(for the team members of startup) (default: {None})
        first_name {STRING} -- the first name of the user that email should be sent (default: {None})
        last_name {STRING} -- the last name of the user that email should be sent (default: {None})
        sub {STRING} -- the subtitle of email (default: {None})
        template {STRING} -- the HTML file that contains email content (default: {None})
        to {STRING} -- the email address of the user that should be send to (default: {None})
        to2 {STRING} -- the email address of the user that should be send to (default: {None})
    
    Returns:
        BOOLEAN -- returns True if email has been sent successfuly
    """
    try:
        message = render_to_string(kwargs.get("template"), {
            'the_message': kwargs.get("the_messages"),
            'the_message2': kwargs.get("the_messages2"),
            'first_name': kwargs.get("first_name"),
            'last_name': kwargs.get("last_name"),
            'domain': kwargs.get("current_site").domain if kwargs.get("current_site") else None,
            'username': kwargs.get("username"),
            'protocol': kwargs.get("protocol"),
            'is_team': kwargs.get("is_team"),
            'image': kwargs.get("image"),
            'to': kwargs.get("to")[0],
            'sub': kwargs.get('sub'),
        })
        subject, from_email, to_the = kwargs.get('sub'), 'support@100startups.ir', kwargs.get('to')
        logger.info(f"email{message}---{to_the}")
        text_content = strip_tags(message)
        msg = EmailMultiAlternatives(
            subject, text_content, from_email, [client for client in kwargs.get('to')] )
        msg.attach_alternative(message, 'text/html')
        msg.send()
        for client in kwargs.get('to'):
            status = f"ایمیل برای {client} ارسال شد"
            the_user = User.objects.filter(email=client).first()
            status_of_user(track_info, status, 'send-email',
                        content1=kwargs.get('the_messages'), content2=kwargs.get('the_messages2'))
            create_notification(the_user, 'send-email', the_user)
    except Exception as e:
        logger.error(str(e))
    return True


def get_list(text):
    """get date to split and use in JalaliDate
    
    Arguments:
        text {STRING} -- the jalali date that been sent by front-end
    
    Returns:
        LIST -- returns list of strings that contains year,month,day in jalai
    """
    lists = []
    for i in text.split('/'):
        lists.append(int(i))
    return lists


def get_list2(text):
    """get date to split and use in JalaliDate
    
    Arguments:
        text {STRING} -- the jalali date that been sent by front-end
    
    Returns:
        LIST -- returns list of strings that contains year,month,day in jalai
    """
    lists = []
    for i in text.split('-'):
        lists.append(int(i))
    return lists




def create_passwd():
    """create randome username and password
    
    Returns:
        STRING -- returns string of a randome username and password 
    """
    username_random = 'u%s' % randint(int('1' * 10), int('9' * 10))
    password = str(uuid.uuid4()).split('-')[-1]
    return username_random, password




def resize_pics(image_b64):
    """decode from base64 if necessary and resize picture to 300x300
    
    Arguments:
        image_b64 {BASE64/PIC} -- base64 encoded image or pure image
    
    Returns:
        IMAGE/FORMAT -- return the content and the format of resized picture
    """
    ext = None
    try:
        format, imgstr = image_b64.split(';base64,')
        ext = format.split('/')[-1]
        data = ContentFile(base64.b64decode(imgstr), name='temp.' + ext)
        image_size = 300.
        img = Image.open(data)
        width, height = img.size
        scale = image_size / width
        img = img.resize((int(width * scale), int(height * scale)), resample=0)
        f = BytesIO()
        try:
            img.save(f, format=ext)
            content = ContentFile(f.getvalue())
        finally:
            f.close()
    except:
        image_size = 300.
        img = Image.open(image_b64)
        width, height = img.size
        scale = image_size / width
        the_format = img.format
        img = img.resize((int(width * scale), int(height * scale)), resample=0)
        f = BytesIO()
        try:
            img.save(f, format=the_format)
            content = ContentFile(f.getvalue())
        finally:
            f.close()
    
    return content, the_format if not ext else ext



def get_date_jalali(year, month, day):
    """convert Jalali date to gregorian
    
    Arguments:
        year {STRING} -- the Jalai year
        month {STRING} -- the jalai month
        day {STRING} -- the jalai day
    
    Returns:
        DATE -- returns converted date from jalai to gregorian
    """
    try:
        date = JalaliDate(year, month, day).to_gregorian()
    except ValueError:
        try:
            date = JalaliDate(year, month, day-1).to_gregorian()
        except ValueError:
            date = JalaliDate(year, month, day-2).to_gregorian()

    return date


def create_notification(user, status, instance):
    """create notificaton records for the specified user on specific object
    
    Arguments:
        user {OBJECT} -- specified user to send notification to
        status {STRING} -- the status for name status in TheStatus table
        instance {OBJECT} -- the object that notification is about
    
    Returns:
        QS -- returns created record if no error happens
    """
    try:
        status = TheStatus.objects.get(name=status)
        return Notification.objects.create_by_model(user=user, status=status, instance=instance)
    except:
        return None



def registerd_startup(function):
    @wraps(function)
    def wrapper(request, *args, **kwargs):
        if request.user.is_authenticated and request.user.user_type == 'startup' and request.user.step == 'finish':
            return redirect('usercp:dashboard_panel')
        return function(request, *args, **kwargs)
    return wrapper